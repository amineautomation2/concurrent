import json
from re import findall
from typing import Literal

import openpyxl
from bs4 import BeautifulSoup
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

from utils import (
    delay,
    fetch_with_backoff,
    find_element_or_none,
    find_elements,
    get_with_backoff,
)


def get_endpoint_by_type(fund_type: Literal["Investment", "ETF"], offset: int):
    if fund_type == "ETF":
        return f"https://www.hl.co.uk/shares/exchange-traded-funds-etfs/list-of-etfs?offset={offset}&etf_search_input=etf&companyid=&sectorid="
    return f"https://www.hl.co.uk/shares/investment-trusts/search-for-investment-trusts?offset={offset}&it_search_input=p&companyid=&sectorid="


def get_funds_url(
    driver: WebDriver, fund_type: Literal["Investment", "ETF"], xlsx_path: str
):
    offset = 0
    max_offset = 0
    page = 1
    pages = 1

    list_funds = []
    endpoint = get_endpoint_by_type(fund_type, offset)
    get_with_backoff(driver, endpoint)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[fund_type]
    print(f"[####] H&L {fund_type} [####]")
    wait = WebDriverWait(driver, 3)

    accept_cookies = find_element_or_none(
        wait, '//*[@id="onetrust-reject-all-handler"]'
    )
    if accept_cookies:
        accept_cookies.click()

    page_list = find_elements(wait, "//table/tbody/tr[1]/td[1]/table/tbody/tr/td/a")
    if page_list:
        max_offset = (len(page_list) + 1) * 50 - 50
        pages = len(page_list) + 1

    while offset <= max_offset:
        # print(f'[#]  H&L [{page}/{pages}]')
        TABLE_XPATH = '//div[@class="table-overflow-wrapper"]/table/tbody'
        ROWS_XPATH = '//div[@class="table-overflow-wrapper"]/table/tbody/tr'
        wait.until(EC.presence_of_element_located((By.XPATH, TABLE_XPATH)))
        fund_rows = find_elements(wait, ROWS_XPATH)
        if fund_rows:
            for fund in fund_rows[1 : len(fund_rows) - 1]:
                name_xpath = "./td[2]/a" if fund_type == "Investment" else "./td[5]"
                url_xpath = "./td[2]/a"
                name = fund.find_element(By.XPATH, name_xpath).text.strip()
                url = fund.find_element(By.XPATH, url_xpath).get_attribute("href")

                list_funds.append(dict(name=name, url=url))
        page += 1
        offset += 50
        delay(2, 3)
        if offset <= max_offset:
            endpoint = get_endpoint_by_type(fund_type, offset)
            get_with_backoff(driver, endpoint)

    iter = 2
    for fund in list_funds:
        ws.cell(iter, 1, fund["name"])
        cell = ws.cell(iter, 3, fund["url"])
        cell.style = "Hyperlink"
        cell.hyperlink = fund["url"]
        iter += 1

    wb.save(xlsx_path)
    wb.close()
    print(f"[#] Parsed {len(list_funds)} funds into {xlsx_path}")


"""
ETF_URL_XPATH = //nav[@aria-label="Factsheet tabs"]/ul/li[3]/div/a
ETF_ISIN_XPATH = //div[@id='radix-:R3km:-content-Overview'][1]/section/div[1]/div[2]/ul/li[last()]/div/div[2]

IT_URL_XPATH = //nav[@aria-label="Factsheet tabs"]/ul/li[6]/div/a
IT_ISIN_XPATH = //div[@id="radix-:r3:-content-Overview"][1]/section/div[1]/div[2]/ul/li[6]/div/div[2]

KEYWORD_XPATH = //div[@id="__next"]/div/div[2]/header/div[3]/div[2]/ul/div/div/div/li
"""


def get_fund_keyword_it(driver: WebDriver, funds: list[dict]) -> list[dict]:
    url_xpath = '//nav[@aria-label="Factsheet tabs"]/ul/li[6]/div/a'
    url2_xpath = '//div[@id="factsheet-nav-container"]/ul/li[8]/a'
    isin_xpath = '//ul[@class="info-list_root__Vpw6y info-list_narrow__gzzia"]'
    keyword_xpath = (
        '//div[@id="__next"]/div/div[2]/header/div[3]/div[2]/ul/div/div/div/li'
    )
    keyword_xpath = '//div[@class="applicable-products_applicable_products__JsXiH"]'
    wait = WebDriverWait(driver, timeout=10)
    data = []
    for fund in funds:
        url_backup = fund.get("url")
        try:
            name = fund["name"]
            isin, url, keyword_fmt = None, None, None
            get_with_backoff(driver, fund["url"])
            accept_cookies = find_element_or_none(
                WebDriverWait(driver, timeout=3),
                '//*[@id="onetrust-reject-all-handler"]',
            )
            if accept_cookies:
                accept_cookies.click()

            url = f"{driver.current_url}/company-information"
            get_with_backoff(driver, url)
            isin = find_element_or_none(wait, isin_xpath)
            if isin:
                res = findall(r"[A-Z]{2}[A-Z0-9]{9}[0-9]", isin.text)
                if len(res) > 0:
                    isin = res[0]
            keyword = find_elements(wait, keyword_xpath)
            if keyword:
                keyword_fmt = []
                for k in keyword:
                    keyword_fmt.append(k.text.strip())
                # keyword_fmt = f"This Stock can be held in a {', '.join(keyword_fmt)}"
                keyword_fmt = f"This stock can be held in a {', '.join(keyword_fmt[: len(keyword_fmt) - 1])} or {keyword_fmt[-1]}"
            f = dict(
                name=name,
                isin=isin,
                url=url or url_backup,
                keyword=keyword_fmt,
                index=fund.get("index"),
                sheet="Investment",
            )

            data.append(f)
        except:
            print(f"error: {fund}")
        # pprint(f)
        delay(1, 3)
    return data


def get_fund_keyword(
    driver: WebDriver, funds: list[dict], fund_type: str
) -> list[dict]:
    # url_xpath = '//nav[@aria-label="Factsheet tabs"]/ul/li[3]/div/a'
    # url2_xpath = '//div[@id="factsheet-nav-container"]/ul/li[5]/a'
    ## isin_xpath = '//div[@id="radix-:R3km:-content-Overview" and @data-state="active"]/section/div[1]/div[2]/ul/li[last()]/div/div[2]'
    ## isin_xpath = '//ul/li/div/div[matches(., "[A-Z]{2}[A-Z0-9]{9}[0-9]")]'
    # isin_xpath = '//ul[@class="info-list_root__Vpw6y info-list_narrow__gzzia"]'
    # keyword_xpath = (
    #    '//*[@id="__next"]/div/div[2]/header/div[3]/div[2]/ul/div/div/div/li'
    # )
    # keyword_xpath = '//*[@id="__next"]/div/div/header/div[3]/div[2]/ul/div/div/div'
    keyword_xpath = '//ul[@class="applicable-products_applicable_products__JsXiH"]'
    # keyword_xpath = '//div[@class="small-hide medium-hide wide-medium-hide"]'
    wait = WebDriverWait(driver, timeout=10)
    data = []
    for fund in funds:
        url_backup = fund.get("url")
        name, isin, url, keyword_fmt = None, None, None, None
        if url_backup:
            try:
                get_with_backoff(driver, url_backup)

                # accept_cookies = find_element_or_none(
                #    WebDriverWait(driver, timeout=3),
                #    '//*[@id="onetrust-reject-all-handler"]',
                # )
                # if accept_cookies:
                #    accept_cookies.click()
                keyword = find_element_or_none(wait, keyword_xpath)
                if keyword:
                    # keyword = keyword.text.replace("\n", ", ")
                    print("Keyword = ", keyword.text)
                    keyword = keyword.text.split("\n")
                    keyword_fmt = f"This stock can be held in a {', '.join(keyword[: len(keyword) - 1])} or {keyword[-1]}"

                # url_etf = "https://www.hl.co.uk/shares/shares-search-results/BF8H5S0"
                next_obj = get_next_object(url_backup)
                if next_obj:
                    query = next_obj.get("query")
                    if query:
                        url = f"https://www.hl.co.uk/shares/shares-search-results/{query['prefix']}/{query['slug']}"
                    props = next_obj.get("props")
                    if props:
                        fund_props = props.get("pageProps")
                        if fund_props:
                            details = fund_props.get("investmentDetails")
                            if details:
                                name = details.get("name")
                                isin = details.get("isin")
                f = dict(
                    name=name,
                    isin=isin,
                    url=url or url_backup,
                    keyword=keyword_fmt,
                    index=fund.get("index"),
                    sheet=fund_type,
                )

                data.append(f)
            except:
                print(
                    f"error: {fund}",
                )
            # pprint(f)
        delay(2, 3)
    return data


def get_next_object(url: str) -> dict | None:
    try:
        # 1. Fetch raw HTML content
        response = fetch_with_backoff(url)
        if response is None:
            print(f"Skipping {url}")
            return None
        # 2. Parse the HTML shell
        soup = BeautifulSoup(response.text, "html.parser")

        # 3. Locate the specific __NEXT_DATA__ script block
        script_tag = soup.find("script", id="__NEXT_DATA__")

        if script_tag:
            # 4. Load the raw string data into a native Python dictionary
            next_data = json.loads(script_tag.string)

            # 5. Extract your target attribute (safely using .get() to prevent crashes)
            return next_data

            # print(f"Source URL: {url} -> Extracted Attribute URL: {target_url}")
        else:
            print(f"data not found on {url}")
            return None

    except Exception as e:
        print(f"An error occurred while scraping {url}: {e}")
        return None
