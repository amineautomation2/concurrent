from pprint import pprint

import requests
import openpyxl
import json
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.support.wait import WebDriverWait
from math import ceil
from utils import setup_driver, delay, find_element_or_none, find_elements, get_xlsx_filepath, get_with_backoff

# TODO: change OptanonAlerboxClosed
cookies = {
    '__losp': 'web_share%3D2-web_index%3D2',
    'OptanonConsent': 'isGpcEnabled=1&datestamp=Wed+Apr+08+2026+10%3A00%3A47+GMT%2B0100+(GMT%2B01%3A00)&version=202503.2.0&browserGpcFlag=1&isIABGlobal=false&hosts=&consentId=cf498ced-bf0f-41b8-b80e-a65774b52610&interactionCount=2&isAnonUser=1&landingPath=NotLandingPage&groups=C0001%3A1%2CC0004%3A0%2CC0002%3A0%2CC0003%3A0&AwaitingReconsent=false&intType=2&geolocation=%3B',
    'LaunchDarklyUser': '01c89fd0-31ab-11f1-a445-97b47562a254',
    'OptanonAlertBoxClosed': '2026-04-06T11:23:20.578Z',
    'mostViewedShares': 'B1YW440',
    'mostViewedFunds': 'BHZCS05%2CB2PB2B6',
}

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:149.0) Gecko/20100101 Firefox/149.0',
    'Accept': 'application/json, text/plain, */*',
    'Accept-Language': 'en-US,en;q=0.9',
    # 'Accept-Encoding': 'gzip, deflate, br, zstd',
    'Sec-GPC': '1',
    'Connection': 'keep-alive',
    'Referer': 'https://www.hl.co.uk/funds/fund-discounts,-prices--and--factsheets/search-results?start=0&rpp=20&lo=0&sort=fd.full_description&sort_dir=asc',
    # 'Cookie': '__losp=web_share%3D2-web_index%3D2; OptanonConsent=isGpcEnabled=1&datestamp=Wed+Apr+08+2026+10%3A00%3A47+GMT%2B0100+(GMT%2B01%3A00)&version=202503.2.0&browserGpcFlag=1&isIABGlobal=false&hosts=&consentId=cf498ced-bf0f-41b8-b80e-a65774b52610&interactionCount=2&isAnonUser=1&landingPath=NotLandingPage&groups=C0001%3A1%2CC0004%3A0%2CC0002%3A0%2CC0003%3A0&AwaitingReconsent=false&intType=2&geolocation=%3B; LaunchDarklyUser=01c89fd0-31ab-11f1-a445-97b47562a254; OptanonAlertBoxClosed=2026-04-06T11:23:20.578Z; mostViewedShares=B1YW440; mostViewedFunds=BHZCS05%2CB2PB2B6',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-origin',
    'Priority': 'u=0',
}


def get_funds_json(headers: dict, cookies: dict, start: int = 0, rpp: int = 20) -> dict:
    endpoint = f'https://www.hl.co.uk/ajax/funds/fund-search/search?investment=&companyid=&sectorid=&wealth=&unitTypePref=&tracker=&payment_frequency=&payment_type=&yield=&standard_ocf=&perf12m=&perf36m=&perf60m=&fund_size=&num_holdings=&start={start}&rpp={rpp}&lo=0&sort=fd.full_description&sort_dir=asc&'
    res = requests.get(endpoint, cookies=cookies, headers=headers)
    return res.json()


def get_funds_url_mf(xlsx_path: str) -> list[dict]:
    funds = []
    start = 0
    rpp = 60
    data = get_funds_json(headers=headers, cookies=cookies)
    total = ceil(data["TotalResults"]/60)

    for i in range(total):
        # print(f"current page [{i+1}/{total}]")
        data = get_funds_json(
            headers=headers, cookies=cookies, start=start, rpp=rpp)
        for fund in data["Results"]:
            name = fund["full_description"]
            sedol = fund["sedol"]
            url = f"https://www.hl.co.uk/funds/fund-discounts,-prices--and--factsheets/search-results/{sedol}"
            funds.append(dict(name=name, url=url))
        start += rpp
    iter = 2
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["MF"]
    for fund in funds:
        ws.cell(iter, 1, fund["name"])
        cell = ws.cell(iter, 3, fund["url"])
        cell.style = "Hyperlink"
        cell.hyperlink = fund["url"]
        iter += 1
    wb.save(xlsx_path)
    wb.close()
    return funds


"""
ETF_URL_XPATH = //nav[@aria-label="Factsheet tabs"]/ul/li[3]/div/a
ETF_ISIN_XPATH = //div[@id='radix-:R3km:-content-Overview'][1]/section/div[1]/div[2]/ul/li[last()]/div/div[2]

IT_URL_XPATH = //nav[@aria-label="Factsheet tabs"]/ul/li[6]/div/a
IT_ISIN_XPATH = //div[@id='radix-:r3:-content-Overview'][1]/section/div[1]/div[2]/ul/li[6]/div/div[2]

KEYWORD_XPATH = //div[@id="__next"]/div/div[2]/header/div[3]/div[2]/ul/div/div/div/li
"""


def get_fund_keyword_mf(driver: WebDriver, funds: list[dict]) -> list[dict]:
    url_xpath = "//div[@id='factsheet-nav-container']/ul/li[contains(., 'Key')]/a"
    isin_xpath = "//h2[@class='tab-divide'][contains(.,'Other')]/following-sibling::div/table/tbody/tr[last()]/td"
    # keyword_xpath = '//div[@class="small-hide medium-hide wide-medium-hide"]'
    keyword_xpath = '/html/body/main/div/div/div/div/div/div[1]/div/div[8]/div[1]'
    wait = WebDriverWait(driver, timeout=5)
    data = []
    for fund in funds:
        url_backup = fund.get("url")
        try:
            name = fund["name"]
            isin, url, keyword = None, None, None
            get_with_backoff(driver, fund['url'])
            url_elm = find_element_or_none(wait, url_xpath)
            if url_elm:
                url = url_elm.get_attribute("href")
                if url:
                    get_with_backoff(driver, url)
                    isin = find_element_or_none(wait, isin_xpath)
                    # print(url)
                    if isin:
                        isin = isin.text
                    keyword = find_element_or_none(wait, keyword_xpath)
                    if keyword:
                        keyword = keyword.text
            f = dict(name=name,
                     isin=isin,
                     url=url or url_backup,
                     keyword=keyword,
                     index=fund.get("index"),
                     sheet="MF",)

            data.append(f)
        except:
            print(f"error: {fund}",)
        # pprint(f)
        delay(0.5, 1)
    return data
