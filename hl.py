from math import ceil
from pprint import pprint
import openpyxl
from mf import get_fund_keyword_mf, get_funds_url_mf
from selenium.webdriver.chrome.webdriver import WebDriver
from utils import delay, setup_driver, get_xlsx_filepath, save_xlsx, get_with_backoff

from worker import (
    add_isin,
    get_xlsx_data,
    get_data_by_worker_id,
    process_data,
    write_csv_by_id,
)


def get_empty_isin(xlsx: str, sheet: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[sheet]
    funds = []
    for i in range(2, ws.max_row + 1):
        if ws.cell(i, 2).value:
            continue
        name = ws.cell(i, 1).value
        url = ws.cell(i, 3).value
        funds.append(dict(name=name, url=url, index=i))
    return funds


def hl_runner(id_worker: int, max_workers: int):
    pref = {
        # "profile.managed_default_content_settings.javascript": 2,
        # "profile.managed_default_content_settings.images": 2,
        # "profile.default_content_setting_values.notifications": 2,
        # "profile.managed_default_content_settings.stylesheets": 2,
    }
    driver = setup_driver(True, pref)
    #    driver.execute_cdp_cmd("Network.setBlockedURLs", {
    #        "urls": [
    #            # "*.google-analytics.com", "*.doubleclick.net", "*.facebook.com",
    #            "*.css", "*.png", "*.jpg", "*.gif", "*.svg", "*.woff*", "*.mp4"
    #        ]
    #    })
    #    driver.execute_cdp_cmd("Network.enable", {})
    #
    # xlsx = get_xlsx_filepath("hl.xlsx")
    # TODO add index to funds
    # get_funds_url(driver, "Investment", xlsx)
    # get_funds_url(driver, "ETF", xlsx)
    # get_funds_url_mf(xlsx)
    # funds = get_empty_isin(xlsx, "ETF")
    # funds = get_empty_isin(xlsx, "Investment")
    funds = get_xlsx_data("spreadsheet/hl.xlsx", "MF")
    funds_per_worker = get_data_by_worker_id(id_worker, max_workers, funds)
    out_csv = f"hl_mf_{id_worker}.csv"
    funds = process_worker_batch(driver, out_csv, funds_per_worker)
    print(len(funds_per_worker))
    driver.quit()


def process_worker_batch(
    driver: WebDriver,
    out_csv: str,
    funds: list[dict],
):
    funds_with_keywords = get_fund_keyword_mf(driver, funds)
    fields = ["index", "name", "isin", "url", "keyword"]
    write_csv_by_id(out_csv, funds_with_keywords, fields)
    print(funds_with_keywords)
    # save_xlsx(xlsx_path=xlsx, funds=funds_with_keywords, cols=cols, sheet="MF")


def process_batch(xlsx: str, driver: WebDriver, funds: list[dict]):
    batch_index = 1
    batch_size = 10
    print(f"Processing Keywords Batch [{batch_index}/{ceil(len(funds) / batch_size)}]")
    for start in range(0, len(funds), batch_size):
        funds_batch = funds[start : start + batch_size]
        if batch_index % 20 == 0:
            print(
                f"Processing Keywords Batch [{batch_index}/{ceil(len(funds) / batch_size)}]"
            )
        funds_with_keywords = get_fund_keyword_mf(driver, funds_batch)
        cols = ["name", "isin", "url", "keyword"]
        save_xlsx(xlsx_path=xlsx, funds=funds_with_keywords, cols=cols, sheet="MF")
        batch_index += 1
    driver.quit()
